# coding=utf-8
import os
import openpyxl
from StudentInfo import StudentInfo


def to_float(num):
    try:
        return float(num)
    except:
        if '+' in unicode(num):
            return 1.0
        return None


def hw_number(task, course_list):
    """
    Возвращает номер домашнего задания
    из порядкового номера задачи
    """
    accumulator = 0
    counter = 1
    for limit in course_list:
        accumulator += limit
        if task < accumulator:
            return counter
        counter += 1
    return None


def AL_parse(file, begin_score_column, StudentInfoList):
    """
    Парсинг файла с оценками по алгоритмам (обе подгруппы)
    @param file: файл в формате OpenXML
    @param begin_score_column: номер столбца (с нуля) где начинаются оценки (у двух подгруппы они разные)
    @param StudentInfoList: общий список студентов
    """

    def build_hw_count(row, start):
        """
        Просматривает выданную строку и заполняет список ALG_LIST - число заданий в каждой домашке
        Заполнение делается в соответствии с правилом: если у новой ячейки есть правая жирная граница, то это конец домашки
        @param row: строка, в которой это правило выполняется
        """
        counter = 0
        for cell in row[start:]:
            style = cell.style.borders.right.border_style
            counter += 1
            if style != 'none':
                ALG_LIST.append(counter)
                counter = 0
        if counter:
            ALG_LIST.append(counter)

    workbook = openpyxl.load_workbook(filename=file)
    for sheet in workbook.worksheets:
        ALG_LIST = [] # число заданий в каждой домашке.
        meta = sheet.rows[0] # column names

        build_hw_count(sheet.rows[0], begin_score_column)

        for row in sheet.rows[1:]:
            if row[0].value is None: continue
            surname, name = row[1].value, row[2].value
            for task, score in enumerate(row[begin_score_column:]):
                task_number = meta[task + begin_score_column].value
                if task_number is None: continue
                student = StudentInfo(name, surname, u"Алгоритмы",
                                 hw_number(int(task), ALG_LIST),
                                 task_number,
                                 to_float(score.value), now)
                StudentInfoList.append(student)


def CO_parse(file, StudentInfoList):
    """
    Парсинг файла с оценками по комбинаторике
    @param file: файл в формате OpenXML
    @param StudentInfoList: общий список студентов
    """

    workbook = openpyxl.load_workbook(filename=file)
    for hw, sheet in enumerate(workbook.worksheets[:-1]):
        meta = sheet.rows[0] # column names
        if hw == 2:
            begin_score_column = 2 # колонка с которой начинаются оценки
            end_score_column = -4 # колонка где заканчиваются оценки
        else:
            begin_score_column = 3
            end_score_column = -1

        for row in sheet.rows[1:]:
            surname, name = row[0].value, row[1].value
            for task, score in enumerate(row[begin_score_column:end_score_column]):
                task_number = meta[task + begin_score_column].value
                if task_number is None: continue
                student = StudentInfo(name, surname, u"Комбинаторика",
                                 int(hw + 1),
                                 task_number,
                                 to_float(score.value), now)
                StudentInfoList.append(student)


def AS_parse(file, StudentInfoList):
    """
    Парсинг файла с оценками по алгебраическим структурам
    @param file: файл в формате OpenXML
    @param StudentInfoList: общий список студентов
    """

    AS_LIST = []

    def build_hw_count(row):
        """
        Просматривает выданную строку и заполняет список AS_LIST - число заданий в каждой домашке
        Заполнение делается в соответствии с правилом: если у новой ячейки есть правая жирная граница, то это конец домашки
        @param row: строка, в которой это правило выполняется
        """
        counter = 0
        for cell in row[3:-8]:
            style = cell.style.borders.right.border_style
            counter += 1
            if style != 'none':
                AS_LIST.append(counter)
                counter = 0
        AS_LIST.append(counter)

    workbook = openpyxl.load_workbook(filename=file)
    for sheet in workbook.worksheets:
        meta = sheet.rows[0] # column names
        build_hw_count(sheet.rows[0])
        for row in sheet.rows[4:]:
            surname, name = row[1].value.split()
            for hw, score in enumerate(row[3:-8]):
                student = StudentInfo(name, surname, u"Алгебраические структуры", hw_number(int(hw), AS_LIST),
                                 meta[hw + 3].value,
                                 to_float(score.value), now)
                StudentInfoList.append(student)


def GT_parse(file, StudentInfoList):
    """
    Парсинг файла с оценками по теории графов
    @param file: файл в формате OpenXML
    @param StudentInfoList: общий список студентов
    """

    GRAPH_LIST = [] # число заданий в каждой домашке.

    def build_hw_count(row):
        """
        Просматривает выданную строку и заполняет список GRAPH_LIST - число заданий в каждой домашке
        Заполнение делается в соответствии с правилом: если у новой ячейки есть правая жирная граница, то это конец домашки
        @param row: строка, в которой это правило выполняется
        """
        counter = 0
        for cell in row[3:-2]:
            style = cell.style.borders.right.border_style
            counter += 1
            if style != 'none':
                GRAPH_LIST.append(counter)
                counter = 0


    workbook = openpyxl.load_workbook(filename=file)
    for sheet in workbook.worksheets:
        meta = sheet.rows[0] # column names
        build_hw_count(sheet.rows[1])
        for row in sheet.rows[1:]:
            surname, name = row[0].value, row[1].value
            for hw, score in enumerate(row[3:-2]):
                student = StudentInfo(name, surname, u"Теория графов", hw_number(int(hw), GRAPH_LIST), meta[hw + 3].value,
                                 to_float(score.value), now)
                StudentInfoList.append(student)


import datetime
now = datetime.datetime.now()

def parse():
    """
    Запуск парсинга всех документов, что указаны в списке courses.py
    @return: список студентов с оценками: список экземпляров класса StudentInfo
    """
    StudentInfoList = []
    folder = "OpenXML"
    ext = "xlsx"

    import courses

    GT_parse("{0}.{1}".format(os.path.join(folder, courses.spreadsheets['GT']), ext), StudentInfoList)
    AS_parse("{0}.{1}".format(os.path.join(folder, courses.spreadsheets['AS']), ext), StudentInfoList)
    CO_parse("{0}.{1}".format(os.path.join(folder, courses.spreadsheets['CO']), ext), StudentInfoList)
    AL_parse("{0}.{1}".format(os.path.join(folder, courses.spreadsheets['AL1']), ext), 5, StudentInfoList)
    AL_parse("{0}.{1}".format(os.path.join(folder, courses.spreadsheets['AL2']), ext), 4, StudentInfoList)

    return StudentInfoList


if __name__ == "__main__":
    list = parse()
    for key, student in enumerate(list):
        print key, student